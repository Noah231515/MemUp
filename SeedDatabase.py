from openpyxl import load_workbook
import sqlite3
import uuid

conn = sqlite3.connect('./MemUp.db')
cursor = conn.cursor()

print("Opening vocab.xlsx")
wb = load_workbook(filename = "./vocab.xlsx")
vocab_sheet = wb.active

words = []
defaultCourseId = "6f770d0a-c3b0-45f0-b78d-2c11f366aaa2"

print("Reading vocab.xlsx and writing sentence values")
for row in vocab_sheet.iter_rows(values_only=True):
    
    wordId = str(uuid.uuid4())
    
    wordData = {
        "id": wordId,
        "vocab_japanese": row[1],
        "vocab_kana": row[2],
        "vocab_furigana": row[3],
        "vocab_english": row[4],
        "pos": row[10]
    }
    
    sentenceData = {
        "sentence_japanese": row[5],
        "sentence_furigana": row[6],
        "sentence_english": row[7],
    }
    
    words.append(wordData)
    
    # Each version of the sentence is treated as a single entry in the table, but all 3 entries are linked to one word, so we insert them as we go through each loop
    # The sentence table is formatted like so: Id, Sentence, WordId, SentenceType 
    for sentence in sentenceData.values():
        sentenceId = uuid.uuid4()
        cursor.execute(f"INSERT INTO Sentence VALUES (?, ?, ?, NULL)", [str(sentenceId), sentence, wordId])

# Create the default course
print("Writing default course values")
cursor.execute("INSERT INTO Courses VALUES (?, 'Japanese Core Vocabulary', 'Core 10,000 vocabulary words for Japanese.')", [defaultCourseId]) 
 
#Write values into Word table and into CourseWord Table
print("Writing Word and CourseWord values")    
for wordData in words:
    # Write each word into the Word Table. The Word Table is formatted like so: Id, VocabJapanese, VocabKana, VocabEnglish, PartOfSpeech
    cursor.execute(f"INSERT INTO Word VALUES (?, ?, ?, ?, ?)", [wordData['id'], wordData['vocab_japanese'], wordData['vocab_kana'], wordData['vocab_english'], wordData['pos']])
    
    # Seed the CourseWord joint table, linking all of our words to the default Japanese Core Vocabulary Course
    cursor.execute("INSERT INTO CourseWord VALUES (?, ?, ?)", [defaultCourseId, wordData['id'], str(uuid.uuid4())])


conn.commit()
conn.close()





